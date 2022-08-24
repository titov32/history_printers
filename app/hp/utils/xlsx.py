import openpyxl
from openpyxl.drawing.image import Image


def gen_cells():
    columns = ['A', 'D', 'G']
    row = 1
    while True:
        for col in columns:
            yield f'{col}{row}', f'{col}{row + 6}', f'{col}{row + 7}',
        row += 8


def create_xlsx_file(response_sql):
    print(response_sql)
    wb = openpyxl.Workbook()
    ws = wb.active
    get_3cell = gen_cells()
    for record in response_sql:
        qr_cell, departament_cell, model_cell = get_3cell.__next__()
        img = Image(record[0].qr)
        img.width = 120
        img.height = 120
        ws.add_image(img, qr_cell)
        ws[departament_cell] = f'{record[0].department.name} ' \
                               f'{record[0].location}'
        ws[model_cell] = f'{record[0].model_printer.brand} ' \
                         f'{record[0].model_printer.model} '
    wb.save('static/xlsx/printers.xlsx')
